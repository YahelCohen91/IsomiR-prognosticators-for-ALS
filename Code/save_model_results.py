import sys,os, re
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import traceback
import csv
import xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook
import math
import matplotlib.patches as mpatches
import scipy.stats
from sksurv.metrics import cumulative_dynamic_auc
from sksurv.metrics import concordance_index_ipcw
from sksurv.nonparametric import kaplan_meier_estimator
from sksurv.functions import StepFunction
import inspect


from sklearn.pipeline import Pipeline
from sklearn.model_selection import GridSearchCV
from sklearn.feature_selection import SelectFromModel
from sklearn.feature_selection import SelectKBest

from mlxtend.feature_selection import SequentialFeatureSelector as SFS
from globals import Usecase_features_selection as features_selection

from models_helpers import my_IBS_inverse
import globals
from globals import excel_file_lock # summary_file_lock, features_selection_file_lock


death_name = 'outcome'
survival_name = 'Survival from onset'
survival_name2 = 'Survival from enrolment'

save_results_file_name = 'best_model_per_loop_{0}.xlsx'
save_results_file_name_coeff_tab_name  = 'best_model_coeff'
save_events_file_name =  'events_{0}.xlsx'
class save_results(object):
    #survival_name = 'Survival from onset'
    #survival_name2 = 'Survival from enrolment'
    DT_token = None
    interval_brier_score_token = r'^brier_m:'
    def __init__(cls, DT_token):
        cls.DT_token = DT_token

    def model(cls, best_model_results, name):
        try:
            # Get the coefficients and hazard ratios
            # Create an Excel workbook and add a worksheet
    
            name_N_path = os.path.join(globals.savefig_path_internal,'model_{0}_params_{1}.xlsx'.format(name, cls.DT_token))


            if os.path.isfile(name_N_path):
                wb = load_workbook(filename=name_N_path)
            else:
                wb = wb = Workbook()
                ws_best_model_coeff = wb.create_sheet('best_model_coeff')

            ws_best_model_coeff = wb['best_model_coeff']


            helper_features = tuple([str(f) for f in best_model_results['features']])

            score_type = best_model_results['score_type'][0]
            ws_best_model_coeff.append((name, 'features') + helper_features)
            ws_best_model_coeff.append((name, 'test_' + score_type) + tuple(best_model_results['mean_score_test']))
            ws_best_model_coeff.append((name, 'train_' + score_type) + tuple(best_model_results['mean_score_train']))
            try:
                ws_best_model_coeff.append((name, 'coeff') + tuple(best_model_results['coeff']))
                ws_best_model_coeff.append((name, 'hazard_ratio') + tuple(best_model_results['hazard_ratio']))

            except KeyError as ex:
                coeff_columns = best_model_results.filter(regex=r'^coeff', axis=1).columns
                hazard_ratio_columns = best_model_results.filter(regex=r'^hazard_ratio', axis=1).columns
                for coeff_name in coeff_columns:
                    ws_best_model_coeff.append((name, coeff_name) + tuple(best_model_results[coeff_name]))
                for hazard_ratio_name in hazard_ratio_columns:
                    ws_best_model_coeff.append((name, hazard_ratio_name) + tuple(best_model_results[hazard_ratio_name]))
            finally:
                wb.save(name_N_path)
                wb.close()

        except Exception as ex:
            print(ex.__context__)
            print(''.join(traceback.format_exception(etype=type(ex), value=ex, tb=ex.__traceback__)))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
            print('bad save_results.perform')

        # @classmethod

    def models_cv(cls, gen_object, param_grid, name):  # purpose - just for the list of runs - not for particlar model params
        try:
            # Get the coefficients and hazard ratios
            # model_trained.named_steps.reduce_dim is pipe['reduce_dim']

            if isinstance(gen_object, GridSearchCV):
                results = pd.DataFrame(gen_object.cv_results_).sort_values(by="mean_test_score", ascending=False)
                results.columns = results.columns.str.replace(r'_test_', '_within_cohort_test_')
            elif isinstance(gen_object, Pipeline):
                print('non relevant for pipelines without gridsearchcv. used only in gridseachcv - when you can see some intermediate steps')
                return

            base = ['mean_within_cohort_test_score', 'rank_within_cohort_test_score','mean_train_score']
            cv_params = list(param_grid.keys())
            cv_params = ['param_'+ c for c in cv_params]
            results = results.drop(columns=['params'])
            results2 = results[cv_params+base]


            # Create an Excel workbook and add a worksheet
            name_N_path = globals.savefig_path + '\save_models_series_cv_results_{0}.xlsx'.format(cls.DT_token)

            if os.path.isfile(name_N_path):
                wb = load_workbook(filename=name_N_path)
            else:
                wb = wb = Workbook()
                ws_cv = wb.create_sheet('cv')
                ws_cv_exparts = wb.create_sheet('cv_exparts')

                l = list(['group']) + results.columns.tolist()
                l2 = list(['group']) + results2.columns.tolist()
                ws_cv.append(l)
                ws_cv_exparts.append(l2)


            # ws = wb.active
            ws_cv = wb['cv']
            ws_cv_exparts = wb['cv_exparts']

            l = list([name]) + results.values.tolist()[0]
            l2 = list([name]) + results2.values.tolist()[0]
            ws_cv.append(l)
            ws_cv_exparts.append(l2)
            '''
            for row in results_tuples:
                ws_cv.append((row))
            for row in results2_tuples:
                ws_cv_exparts.append((row))
            '''
            wb.save(name_N_path)
            wb.close()


        except Exception as ex:
            print(ex.__context__)
            print(''.join(traceback.format_exception(etype=type(ex), value=ex, tb=ex.__traceback__)))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
            print('bad save_results.models_cv')




    def predcition_vs_gt(cls, features, y, trained_model, name=''):  #was standalone save_predcition_vs_gt
        try:

            predictions = trained_model.predict(features)
            patient_ids = list(features.index.values.tolist())
            survival_gt = y[cls.survival_name2]

            # Combine predictions and true survival times into a DataFrame
            comparison_df = pd.DataFrame({
                'Patient ID': patient_ids,
                'True Survival Time': survival_gt,
                'Predicted Survival Time': predictions
            })

            # Save the comparison DataFrame to a CSV file
            comparison_df.to_csv(globals.savefig_path + '\model_{0}_prediction_comparison_{1}.csv'.format(name, cls.DT_token),
                                 index=True)

        except Exception as ex:
            print(ex.__context__)
            print(''.join(traceback.format_exception(etype=type(ex), value=ex, tb=ex.__traceback__)))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
            print('bad save_model_results.predcition_vs_gt')

    def various_predictions_probabilities_per_patient(cls, features, y, trained_model, prediction_vector_times, name=''):
        try:

            if isinstance(trained_model, SFS):
                trained_model = trained_model.estimator  # move from wrapper (SFS/gridsearchcv/pipeline to model)

            # Predict survival function for patients

            survival_gt_time = y[survival_name2]
            survival_gt_death = y[death_name]


            model_surv_prob = np.row_stack([fn(prediction_vector_times) for fn in trained_model.predict_survival_function(features)])



            median_survival_month =  np.full((len(model_surv_prob)), 'NA', dtype=object) #     -1 * np.ones(len(model_surv_prob))
            for row_index, row in enumerate(model_surv_prob):
                # Find the index of the element in the row that corresponds to the mean survival value
                try:

                    logic_occurs = np.where(row <= 0.5)
                    if len(logic_occurs) == 1 and len(logic_occurs[0]) == 0:
                        print(f'in {name} median_survival_month: all survival vector greater than median chance in row # {row_index}')
                        continue
                    index_in_row = logic_occurs[0][0]
                    if index_in_row & index_in_row > 0:
                        prob_after = row[index_in_row]
                        prob_before = row[index_in_row-1]
                        mechane = (prob_before - prob_after)
                        if mechane:
                            fraction_lean =  - ((0.5 - prob_after) / mechane)  # bilinear_interpolation
                        else:
                            fraction_lean =  - 0.5
                        median_survival_month[row_index] = prediction_vector_times[index_in_row] + fraction_lean
                    else:
                        print(f'median_survival_month: less than mimimal ruler value of {prediction_vector_times[0]} in {index_in_row}')
                        continue
                except Exception as ex:
                    print(f'bad median survival time for row_index:{row_index}, row:{row}')
                    print(ex.__context__)
                    print(''.join(traceback.format_exception(etype=type(ex), value=ex, tb=ex.__traceback__)))
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print(exc_type, fname, exc_tb.tb_lineno)



            # Calculate survival probability at specific time points
            time_points = list(prediction_vector_times)  
            survival_probabilities =  np.row_stack([fn(time_points) for fn in trained_model.predict_survival_function(features)])
            survival_probabilities_at_event = list(-1 * np.ones(len(features)))
            survival_probabilities_diff_at_death_or_censorship = list(-1 * np.ones(len(features)))
            for i, t in enumerate(survival_gt_time):
                survival_probabilities_at_event[i] = [fn(t) for fn in trained_model.predict_survival_function(features.iloc[[i]])][0]
                survival_probabilities_diff_at_death_or_censorship[i] = (survival_probabilities_at_event[i]) if survival_gt_death[i] else (1-survival_probabilities_at_event[i])

            survival_probabilities = np.array(survival_probabilities).reshape(len(model_surv_prob), len(time_points))


            # Calculate risk scores
            risk_scores = trained_model.predict(features) 

            patient_ids = list(features.index.values.tolist())


            # Combine predictions and true survival times into a DataFrame
            results_df = pd.DataFrame({
                'Patient ID': patient_ids,
                'True Survival Time': survival_gt_time,
                'death': survival_gt_death,
                'Survival Chance @ Event': survival_probabilities_at_event,
                'survival probabilities diff @ event(death/censorship)': survival_probabilities_diff_at_death_or_censorship,
                'month of median Survival Chance': median_survival_month,
                'RiskScore':risk_scores
            })
            #add also km prediction per pateint

            for i,t in enumerate(time_points):
                results_df[f'SurvivalProbability@ {t} Months'] = survival_probabilities[:,i]

            # Save the comparison DataFrame to a CSV file            
            results_df.to_csv(os.path.join(globals.savefig_path_internal, 'model_{0}_various_predictions_probabilities_per_patient_{1}.csv'.format(name, cls.DT_token)),
                                     index=True)

        except Exception as ex:
            print(ex.__context__)
            print(''.join(traceback.format_exception(etype=type(ex), value=ex, tb=ex.__traceback__)))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
            print('bad save_model_results.various_predictions_probabilities_per_patient')
            
    def summary(cls,  score_results, score_results_vectors, params, input_features_names, y_train, y_test, feature_test,
                feature_train, model_trained, name, use_different_score_metrics,  optimization, prediction_vector_times,
                prediction_vector_times_test = None):

        try:

            if isinstance(model_trained, SFS):
                model_trained = model_trained.estimator  # move from wrapper (SFS/gridsearchcv/pipeline to model)

            if use_different_score_metrics:

                results_metrics = pd.DataFrame({
                    'name': name,
                    'model_mean_test': score_results['mean_score_test'],
                    'model_mean_train': score_results['mean_score_train'],
                    'method': score_results['score_type']


                })
                if prediction_vector_times_test is None:
                    time_interval = prediction_vector_times
                else:
                    time_interval_train = prediction_vector_times
                    time_interval_test = prediction_vector_times_test
                    
                model_test = score_results_vectors['median_brier_test_score']
                model_train = score_results_vectors['median_brier_train_score']

            else:
                min = 8
                max = 74 

                ### ### ### ### ### ### ### ### ### ### ### ### ### ### ### ### 
                # CHANGE HERE FOR DIFFERENT TIME POINTS FOR SURVIVAL FUNCTION #
                ### ### ### ### ### ### ### ### ### ### ### ### ### ### ### ### 
                
                time_interval = np.arange(min, max, 2)  #0.2

                model_risk_scores_test = model_trained.predict(feature_test)
                model_test, model_mean_test = cumulative_dynamic_auc(y_train, y_test, model_risk_scores_test,
                                                                       time_interval)  # y_train - needed for the death/censors events

                model_risk_scores_train = model_trained.predict(feature_train)
                model_train, model_mean_train = cumulative_dynamic_auc(y_train, y_train, model_risk_scores_train,
                                                                             time_interval)

                results_metrics = pd.DataFrame({
                    'name': name,
                    'time_interval': [time_interval],
                    'model_test': [model_test],
                    'model_train': [model_train],
                    'model_mean_test': model_mean_test,
                    'model_mean_train': model_mean_train,
                    'method': 'auc'
                })


            # Create an Excel workbook and add a worksheet
            with excel_file_lock:
                name_N_path = os.path.join(globals.savefig_path, save_results_file_name.format(cls.DT_token))
                if use_different_score_metrics:
                    method = score_results['score_type'][0]
                else:
                    method = 'AUC-t'

                if os.path.isfile(name_N_path):
                    wb = load_workbook(filename = name_N_path)
                else:
                    # create sheets for "best_model_per_loop" excel file
                    wb = Workbook()
                    empty_sheet = wb['Sheet']
                    wb.remove(empty_sheet)
                    ws_test = wb.create_sheet(method + ' test')
                    ws_train = wb.create_sheet(method + ' train')
                    ws_norm_test = wb.create_sheet(method + ' norm_test')
                    ws_norm_train = wb.create_sheet(method + ' norm_train')
                    ws_summary = wb.create_sheet(method)  # 'summary_' - now in the name
                    ws_best_model_coeff = wb.create_sheet('best_model_coeff')
                    ws_params = wb.create_sheet('params')
                    
                    # add headers for columns names in sheets
                    if prediction_vector_times_test is None:
                        ws_test.append(('group','mode')+tuple(time_interval))
                        ws_norm_train.append(('group','mode')+tuple(time_interval))
                        ws_norm_test.append(('group', 'mode') + tuple(time_interval))
                        ws_train.append(('group', 'mode') + tuple(time_interval))
                    else:
                        ws_test.append(('group','mode')+tuple(time_interval_test))
                        ws_norm_train.append(('group','mode')+tuple(time_interval_train))
                        ws_norm_test.append(('group', 'mode') + tuple(time_interval_test))
                        ws_train.append(('group', 'mode') + tuple(time_interval_train))
                    
                    # if ImBS is used
                    if use_different_score_metrics:
                        helper = ('group', 'IBS_test', 'IBS_train', 'ImBS_test', 'ImBS_train', 'score_complement_ImBS_test','score_complement_ImBS_train', 'optimization', '#f')
                        pattern = re.compile(cls.interval_brier_score_token)
                        brier_interval_scores_col = [key for key in score_results_vectors.keys() if pattern.match(key)]
                        #brier_interval_scores_col = score_results_vectors.filter(regex=cls.interval_brier_score_token, axis=1).columns
                        for c in brier_interval_scores_col:
                            helper = helper + (c,)
                        helper = helper + ('features',)
                        ws_summary.append(helper)
                    else:
                        ws_summary.append(('group', 'auct_test', 'auct_train', 'ci_test', 'ci_train','#f' ,'features'))

                #ws = wb.active
                ws_test = wb[method + ' test']
                ws_train = wb[method + ' train']
                ws_norm_test = wb[method + ' norm_test']
                ws_norm_train = wb[method + ' norm_train']
                ws_summary = wb[method]   #'summary_'+
                ws_best_model_coeff =wb['best_model_coeff']
                ws_params = wb['params']


                # Write coefficients and hazard ratios to the worksheet
                ws_test.append(((name,'test',)+tuple(model_test)))
                ws_train.append(((name,'train',)+tuple(model_train)))
                ws_norm_test.append(((name, 'norm_test',) + tuple(model_test)))
                ws_norm_train.append(((name, 'norm_train',) + tuple(model_train)))
                                
                helper_features = tuple([ str(f) for f in score_results['features']])
                selected_features_num = len(helper_features)
                helper_all_features = tuple([str(f) for f in input_features_names])

                if use_different_score_metrics:
                    score_complement_ImBS_train = 1 - np.sqrt(score_results_vectors['ImBS_train'])
                    score_complement_ImBS_test = 1 - np.sqrt(score_results_vectors['ImBS_test'])
                    helper_tuple = (name, score_results['mean_score_test'][0], score_results['mean_score_train'][0], score_results_vectors['ImBS_test'], score_results_vectors['ImBS_train'],
                                    score_complement_ImBS_test,  score_complement_ImBS_train, optimization, str(selected_features_num)  )
                    pattern = re.compile(cls.interval_brier_score_token)
                    brier_interval_scores_col = [key for key in score_results_vectors.keys() if pattern.match(key)]
                    for c in brier_interval_scores_col:
                        helper_tuple = helper_tuple + (score_results_vectors[c],)
                    helper_tuple = helper_tuple + helper_features
                else:
                    helper_tuple = (name, model_mean_test, model_mean_train, score_results['mean_score_test'][0], score_results['mean_score_train'][0], str(selected_features_num) )  + helper_features 
                    print('\n\n',helper_tuple,'\n\n')
                ws_summary.append(helper_tuple)
                ws_best_model_coeff.append((name, ' input_feature_names') + helper_all_features)
                ws_best_model_coeff.append((name, 'features') + helper_features)
                try:
                    ws_best_model_coeff.append((name, 'coeff') + tuple(score_results['coeff']))
                    ws_best_model_coeff.append((name, 'hazard_ratio') + tuple(score_results['hazard_ratio']))
                except KeyError as ex:
                    ws_best_model_coeff.append((name, 'coeff')+ tuple(['multiple coeff due to multiple alpha penalty coeff']))

                ws_params.append(str(l) for l in list(params.items()))
                wb.save(name_N_path)       #So pandas to_excel is thread safe while openpyxl and xlsxwriter are not - this is true even if you write to different excel files
                wb.close()

                return results_metrics

        except Exception as ex:
            print(ex.__context__)
            print(''.join(traceback.format_exception(etype=type(ex), value=ex, tb=ex.__traceback__)))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
            print('bad summary')

        # @classmethod

    def features_selection(cls,selector, features_selection, name, name2, scoring_family, use_different_score_metrics = True):
        try:
            intermediate_steps_log = pd.DataFrame()

            if isinstance(selector, SelectFromModel) or isinstance(selector, SelectKBest):
                score_func_name = selector.get_params()['score_func'].__name__
                if globals.Usecase_features_selection.Stats.value == features_selection.value or globals.Usecase_features_selection.Survival_Univariate_cox.value == features_selection.value:
                    best_score = score_func_name+str(selector.scores_)
                else:
                    best_score = selector.threshold_
                num_f_selected = sum(selector.get_support())
                best_indeces = 'SelectFromModel/SelectKBest. #f: {0}'.format(num_f_selected)
                best_f_names = selector.get_feature_names_out()

            if isinstance(selector, SFS):
                try:
                    best_indeces = selector.k_feature_idx_
                    best_f_names = selector.k_feature_names_

                    best_score = selector.k_score_
                    metric_dict = selector.get_metric_dict(confidence_interval=0.95)
                    if metric_dict:
                        intermediate_steps_log = pd.DataFrame.from_dict(metric_dict).T  # convert selector.subsets_
                        intermediate_steps_log = intermediate_steps_log.astype(str)
                    else:
                        print('selector metric_dict(confidence_interval=0.95) non exist. some log detalies skipped')
                except RuntimeWarning as warning:
                    print('selector metric_dict(confidence_interval=0.95) non exist or other warnning. some log detalies skipped')





            results_best = pd.DataFrame({
                'name': name,
                'best_score': best_score,
                'best_f_names': [best_f_names],
                'best_indeces': [best_indeces],
                'scoring_family':scoring_family,
            })

            if use_different_score_metrics:
                #IBS = my_IBS_inverse(best_score)    #my_IBS_inverse can be used also for reverse ImBS
                #results_best['IBS'] = IBS
                ImBS = my_IBS_inverse(best_score)  # my_IBS_inverse can be used also for reverse ImBS
                results_best['ImBS'] = ImBS

            # Create an Excel workbook and add a worksheet
            #with features_selection_file_lock:
            with excel_file_lock:
                try:
                    name_N_path = os.path.join(globals.savefig_path,'{1}_features_selection_params_{0}.xlsx'.format(cls.DT_token, name2))

                    if os.path.isfile(name_N_path):
                        wb = load_workbook(filename=name_N_path)

                    else:
                        wb = Workbook()
                        ws_best_preproc = wb.create_sheet('best_preproc')
                        ws_steps_log = wb.create_sheet('steps_log')
                        l = list(['group']) + results_best.columns.tolist()
                        l2 = list(['group']) + intermediate_steps_log.columns.tolist()
                        ws_best_preproc.append(l)
                        ws_steps_log.append(l2)


                    ws_best_preproc = wb['best_preproc']
                    ws_steps_log = wb['steps_log']


                    l = list([name]) + results_best.values.tolist()[0]
                    l[3] = str(l[3])
                    l[4] = str(l[4])
                    ws_best_preproc.append(l)

                    #list([name])
                    ll = intermediate_steps_log.values.tolist()
                    for l2 in ll:
                        ws_steps_log.append(list([name])+l2)
                        
                except Exception as ex:
                    print(ex.__context__)
                    print(''.join(traceback.format_exception(type=type(ex), value=ex, tb=ex.__traceback__)))
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    current_function_name = inspect.currentframe().f_back.f_code.co_name
                    print('inside of:')
                    print(current_function_name)
                
                finally:
                    wb.save(name_N_path)
                    wb.close()

        except Exception as ex:
            print(ex.__context__)
            print(''.join(traceback.format_exception(type=type(ex), value=ex, tb=ex.__traceback__)))   # was etype=type(ex)
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            current_function_name = inspect.currentframe().f_back.f_code.co_name
            print(current_function_name)

        