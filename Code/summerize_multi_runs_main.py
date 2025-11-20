import sys,os
import numpy as np
import ast
from collections import defaultdict
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import pandas as pd
import traceback
import datetime
import re
import inspect
import warnings
import mlflow
import mlflow.sklearn
import globals
from globals import os_name

from save_model_results import save_results_file_name, save_results_file_name_coeff_tab_name
from globals import root   
from globals import folder_path, folder_path_internal, file_prefix, multiple_runs_stats_file_name,f_w_tab_name, m_w_tab_name, m_f_inc_tab_name   # token
#from isomiRs_pipelines_main import f_repetition_num
from globals import summary_tab_name, summary_tab_name_ci 

errors_tab_name = 'errors'


def summerize(token, use_different_score_metrics = True):
    try:
        united_summary_df = []
        missing_groups = []
        if use_different_score_metrics:
            print(folder_path)
            summary_file_name = os.path.join(root, folder_path, save_results_file_name.format(token))
            # Read the Excel file into a DataFrame
            summary_df = pd.read_excel(summary_file_name, sheet_name=summary_tab_name)
            verifcation_summary_text = 'summary file # records:{}'.format(len(summary_df))

            # Extract 'group' and 'features' columns
            summary_df = summary_df[['group', 'ImBS_test', 'ImBS_train', '#f']]
            for i, g in enumerate(summary_df['group']):
                try:
                    print(f'group#{i}')
                    group_run_file_name = os.path.join(root, folder_path_internal.format(token),(file_prefix + '_' + g + '_' + token + '.xlsx'))

                    start_position = group_run_file_name.find('group')
                    start_position2 = group_run_file_name.find('_oo')
                    # Check if the prefix is found
                    if start_position != -1:
                        # Extract the substring starting from the prefix
                        group_num = group_run_file_name[start_position+len('group'):start_position2]
                    
                    group_run_df = pd.read_excel(group_run_file_name)
                    group_run_df['ImBS_test'] = summary_df.at[i, 'ImBS_test']
                    group_run_df['ImBS_train'] = summary_df.at[i, 'ImBS_train']
                    group_run_df['#f'] = summary_df.at[i, '#f']
                    group_run_df['group'] = group_num

                    is_numerical_coeff = group_run_df['feature_importances_'] != 'not_selected'
                    group_run_df.loc[~is_numerical_coeff, 'ImBS_test'] = 1.0  # when f was not selected his mark is 0 i.e. max loss and not the mark other features got
                    group_run_df.loc[~is_numerical_coeff, 'ImBS_train'] = 1.0  # when f was not selected his mark is 0 i.e. max loss and not the mark other features got
                    sum_coeff = group_run_df[is_numerical_coeff]['feature_importances_'].abs().sum()
                    if sum_coeff >0:
                        group_run_df.loc[is_numerical_coeff, 'feature_importances_L1_perc'] =  100 * group_run_df.loc[ is_numerical_coeff, 'feature_importances_'] / sum_coeff
                    else:
                        group_run_df.loc[is_numerical_coeff, 'feature_importances_L1_perc'] = None
                    united_summary_df.append(group_run_df)
                except OSError as e:
                    print(f"Unable to open {group_run_file_name}: {e}", file=sys.stderr)
                    missing_groups.append(group_run_file_name)
                except FileNotFoundError as e:
                    print(f"file not found {group_run_file_name}: {e}", file=sys.stderr)
                    missing_groups.append(group_run_file_name)
                except Exception as e:
                    # Handle other exceptions
                    print(f"An error occurred: {e} on group {group_run_file_name}")
                    missing_groups.append(group_run_file_name)
                    
            # Concatenate DataFrames
            united_summary_df = pd.concat(united_summary_df, ignore_index=True)

            unique_f_names = united_summary_df['features'].unique()
            unique_g_names = united_summary_df['group'].unique()

            verifcation_summary_text += '# unique groups on unified:{}'.format(len(unique_g_names))
            verifcation_summary_text += '# unique f:{}'.format(len(unique_f_names))
            verifcation_summary_text += '# missing f:{}'.format(len(missing_groups))

            # Count the number of times each feature appears in each DataFrame
            counts_per_df = united_summary_df.groupby(['features']).size()
            counts_per_df = counts_per_df.to_frame(name='counts_per_feature')

            agg_ImBS_test = united_summary_df.groupby(['features'])['ImBS_test'].agg(lambda x: list(x)).reset_index()
            agg_ImBS_train = united_summary_df.groupby(['features'])['ImBS_train'].agg(lambda x: list(x)).reset_index()
            agg_num_f = united_summary_df.groupby(['features'])['#f'].agg(lambda x: list(x)).reset_index()
            agg_feature_importances_L1_perc = united_summary_df.groupby(['features'])['feature_importances_L1_perc'].agg(lambda x: list(x)).reset_index()
            agg_feature_importances_ = united_summary_df.groupby(['features'])['feature_importances_'].agg( lambda x: list(x)).reset_index()
            agg_group = united_summary_df.groupby(['features'])['group'].agg(  lambda x: list(x)).reset_index()

            agg_ImBS_test.set_index('features', inplace=True)
            agg_ImBS_train.set_index('features', inplace=True)
            agg_num_f.set_index('features', inplace=True)
            agg_feature_importances_L1_perc.set_index('features', inplace=True)
            agg_feature_importances_.set_index('features', inplace=True)
            agg_group.set_index('features', inplace=True)

            agg_vectors = agg_ImBS_test.join(agg_ImBS_train, lsuffix='_left', rsuffix='_right', how='outer')
            agg_vectors = agg_vectors.join(agg_num_f, lsuffix='_left', rsuffix='_right', how='outer')
            agg_vectors = agg_vectors.join(agg_feature_importances_L1_perc, lsuffix='_left', rsuffix='_right', how='outer')
            agg_vectors = agg_vectors.join(agg_feature_importances_, lsuffix='_left', rsuffix='_right', how='outer')
            agg_vectors = agg_vectors.join(agg_group, lsuffix='_left', rsuffix='_right', how='outer')
            agg_vectors.rename(columns={'ImBS_test': 'all_ImBS_test', 'ImBS_train': 'all_ImBS_train', 'feature_importances_L1_perc': 'all_feature_importances_L1_perc',
                                        'feature_importances_': 'all_feature_importances_', '#f':'all_#f'}, inplace=True)

            selected_df = united_summary_df[united_summary_df['feature_importances_'] != 'not_selected']
            not_selected_df = united_summary_df[united_summary_df['feature_importances_'] == 'not_selected']

            selected_counts = selected_df.groupby('features').size()
            selected_counts = selected_counts.to_frame(name='selected_counts')
            not_selected_counts = not_selected_df.groupby('features').size()
            not_selected_counts = not_selected_counts.to_frame(name='not_selected_counts').astype(int)

            features_summary_stats  = selected_df.groupby('features')[['feature_importances_L1_perc']].agg(['median', 'std'])
            features_summary_stats.columns = [f'{col[0]}_{col[1]}' for col in features_summary_stats.columns]

            features_summary_stats2 = united_summary_df.groupby('features')[['ImBS_test']].agg(['median', 'std','min'])  
            features_summary_stats2.columns = [f'{col[0]}_{col[1]}' for col in features_summary_stats2.columns]
            features_summary_stats3 = united_summary_df.groupby('features')[['ImBS_train']].agg( ['median', 'std','min']) 
            features_summary_stats3.columns = [f'{col[0]}_{col[1]}' for col in features_summary_stats3.columns]
            features_summary_stats4 = selected_df.groupby('features')[['#f']].agg(['median', 'std'])
            features_summary_stats4.columns = [f'{col[0]}_{col[1]}' for col in features_summary_stats4.columns]
            features_summary_stats = features_summary_stats.join(features_summary_stats2, lsuffix='_left', rsuffix='_right', how='outer')
            features_summary_stats = features_summary_stats.join(features_summary_stats3, lsuffix='_left', rsuffix='_right', how='outer')
            features_summary_stats = features_summary_stats.join(features_summary_stats4, lsuffix='_left', rsuffix='_right',  how='outer')


            #adding from summary_df model run score
            stats_summary_df = features_summary_stats.join(selected_counts, lsuffix='_left', rsuffix='_right', how='outer')
            stats_summary_df = stats_summary_df.join(not_selected_counts, lsuffix='_left', rsuffix='_right', how='outer')
            stats_summary_df['not_selected_counts'] = stats_summary_df['not_selected_counts'].fillna(0).astype(int)
            stats_summary_df = stats_summary_df.join(counts_per_df, lsuffix='_left', rsuffix='_right', how='outer')
            stats_summary_df = stats_summary_df.join(agg_vectors, lsuffix='_left', rsuffix='_right', how='outer')
            stats_summary_df['selected_counts'] = stats_summary_df['selected_counts'].fillna(0)
            stats_summary_df[f'% selected'] = np.where(stats_summary_df['counts_per_feature'] != 0,
                                            (stats_summary_df['selected_counts'] / stats_summary_df['counts_per_feature']) * 100, 0).astype(int)
            # Set display options to show all columns and rows without truncation
            pd.set_option('display.max_columns', None)
            pd.set_option('display.max_rows', None)
            print(verifcation_summary_text)

            name_N_path_multiple_runs_stats = os.path.join(globals.savefig_path, multiple_runs_stats_file_name.format(token))
            stats_summary_df.to_excel(name_N_path_multiple_runs_stats, index=True, sheet_name=summary_tab_name, engine='xlsxwriter')
            print(f'writing file:{name_N_path_multiple_runs_stats}')
            workbook = load_workbook(name_N_path_multiple_runs_stats)
            errors_sheet = workbook.create_sheet(errors_tab_name)
            next_row = errors_sheet.max_row + 1
            column_number = 1  # Replace with the desired column number
            helper = f'input file to collect groups names is:{group_run_file_name}'
            errors_sheet.cell(row=next_row, column=column_number, value=helper)
            next_row = errors_sheet.max_row + 1
            errors_sheet.cell(row=next_row, column=column_number, value=verifcation_summary_text)
            next_row = errors_sheet.max_row + 2
            errors_sheet.cell(row=next_row, column=column_number, value='missing groups are:')
            for i in missing_groups:
                errors_sheet.cell(row=errors_sheet.max_row + 1, column=column_number, value=i)
                
            workbook.save(name_N_path_multiple_runs_stats)
            workbook.close()
        else:
            print(folder_path)
            summary_file_name = os.path.join(root, folder_path, save_results_file_name.format(token))
            # Read the Excel file into a DataFrame
            summary_df = pd.read_excel(summary_file_name, sheet_name=summary_tab_name_ci)
            verifcation_summary_text = 'summary file # records:{}'.format(len(summary_df))

            # Extract 'group' and 'features' columns
            summary_df = summary_df[['group', 'ci_test', 'ci_train','auct_test','auct_train','#f']]
            for i, g in enumerate(summary_df['group']):
                try:
                    print(f'group#{i}')
                    group_run_file_name = os.path.join(root, folder_path_internal.format(token),(file_prefix + '_' + g + '_' + token + '.xlsx'))

                    start_position = group_run_file_name.find('group')
                    start_position2 = group_run_file_name.find('_oo')
                    # Check if the prefix is found
                    if start_position != -1:
                        # Extract the substring starting from the prefix
                        group_num = group_run_file_name[start_position+len('group'):start_position2]
                    
                    group_run_df = pd.read_excel(group_run_file_name)
                    group_run_df['ci_test'] = summary_df.at[i, 'ci_test']
                    group_run_df['ci_train'] = summary_df.at[i, 'ci_train']
                    group_run_df['auct_test'] = summary_df.at[i, 'auct_test']
                    group_run_df['auct_train'] = summary_df.at[i, 'auct_train']
                    group_run_df['#f'] = summary_df.at[i, '#f']
                    group_run_df['group'] = group_num

                    is_numerical_coeff = group_run_df['feature_importances_'] != 'not_selected'
                    group_run_df.loc[~is_numerical_coeff, 'ci_test'] = 0.5  # when f was not selected his mark is 0.5
                    group_run_df.loc[~is_numerical_coeff, 'ci_train'] = 0.5  # when f was not selected his mark is 0.5
                    group_run_df.loc[~is_numerical_coeff, 'auct_test'] = 0.5  # when f was not selected his mark is 0.5
                    group_run_df.loc[~is_numerical_coeff, 'auct_train'] = 0.5  # when f was not selected his mark is 0.5
                    sum_coeff = group_run_df[is_numerical_coeff]['feature_importances_'].abs().sum()
                    if sum_coeff >0:
                        group_run_df.loc[is_numerical_coeff, 'feature_importances_L1_perc'] =  100 * group_run_df.loc[ is_numerical_coeff, 'feature_importances_'] / sum_coeff
                    else:
                        group_run_df.loc[is_numerical_coeff, 'feature_importances_L1_perc'] = None
                    united_summary_df.append(group_run_df)
                except OSError as e:
                    print(f"Unable to open {group_run_file_name}: {e}", file=sys.stderr)
                    missing_groups.append(group_run_file_name)
                except FileNotFoundError as e:
                    print(f"file not found {group_run_file_name}: {e}", file=sys.stderr)
                    missing_groups.append(group_run_file_name)
                except Exception as e:
                    # Handle other exceptions
                    print(f"An error occurred: {e} on group {group_run_file_name}")
                    missing_groups.append(group_run_file_name)
                    
            # Concatenate DataFrames
            united_summary_df = pd.concat(united_summary_df, ignore_index=True)

            unique_f_names = united_summary_df['features'].unique()
            unique_g_names = united_summary_df['group'].unique()

            verifcation_summary_text += '# unique groups on unified:{}'.format(len(unique_g_names))
            verifcation_summary_text += '# unique f:{}'.format(len(unique_f_names))
            verifcation_summary_text += '# missing f:{}'.format(len(missing_groups))

            # Count the number of times each feature appears in each DataFrame
            counts_per_df = united_summary_df.groupby(['features']).size()
            counts_per_df = counts_per_df.to_frame(name='counts_per_feature')

            agg_ci_test = united_summary_df.groupby(['features'])['ci_test'].agg(lambda x: list(x)).reset_index()
            agg_ci_train = united_summary_df.groupby(['features'])['ci_train'].agg(lambda x: list(x)).reset_index()
            agg_auc_test = united_summary_df.groupby(['features'])['auct_test'].agg(lambda x: list(x)).reset_index()
            agg_auc_train = united_summary_df.groupby(['features'])['auct_train'].agg(lambda x: list(x)).reset_index()
            agg_num_f = united_summary_df.groupby(['features'])['#f'].agg(lambda x: list(x)).reset_index()
            agg_feature_importances_L1_perc = united_summary_df.groupby(['features'])['feature_importances_L1_perc'].agg(lambda x: list(x)).reset_index()
            agg_feature_importances_ = united_summary_df.groupby(['features'])['feature_importances_'].agg( lambda x: list(x)).reset_index()
            agg_group = united_summary_df.groupby(['features'])['group'].agg(  lambda x: list(x)).reset_index()

            agg_ci_test.set_index('features', inplace=True)
            agg_ci_train.set_index('features', inplace=True)
            agg_auc_test.set_index('features', inplace=True)
            agg_auc_train.set_index('features', inplace=True)
            agg_num_f.set_index('features', inplace=True)
            agg_feature_importances_L1_perc.set_index('features', inplace=True)
            agg_feature_importances_.set_index('features', inplace=True)
            agg_group.set_index('features', inplace=True)

            agg_vectors = agg_ci_test.join(agg_ci_train, lsuffix='_left', rsuffix='_right', how='outer')
            agg_vectors = agg_vectors.join(agg_auc_test, lsuffix='_left', rsuffix='_right', how='outer')
            agg_vectors = agg_vectors.join(agg_auc_train, lsuffix='_left', rsuffix='_right', how='outer')
            agg_vectors = agg_vectors.join(agg_num_f, lsuffix='_left', rsuffix='_right', how='outer')
            agg_vectors = agg_vectors.join(agg_feature_importances_L1_perc, lsuffix='_left', rsuffix='_right', how='outer')
            agg_vectors = agg_vectors.join(agg_feature_importances_, lsuffix='_left', rsuffix='_right', how='outer')
            agg_vectors = agg_vectors.join(agg_group, lsuffix='_left', rsuffix='_right', how='outer')
            agg_vectors.rename(columns={'ci_test': 'all_ci_test', 'ci_train': 'all_ci_train',
                                        'auct_test' : 'all_auc_time_test', 'auct_train' : 'all_auc_time_train',
                                        'feature_importances_L1_perc': 'all_feature_importances_L1_perc',
                                        'feature_importances_': 'all_feature_importances_', '#f':'all_#f'}, inplace=True)
            
            # Filter non-'not_selected' rows for calculating median and std
            selected_df = united_summary_df[united_summary_df['feature_importances_'] != 'not_selected']
            not_selected_df = united_summary_df[united_summary_df['feature_importances_'] == 'not_selected']

            selected_counts = selected_df.groupby('features').size()
            selected_counts = selected_counts.to_frame(name='selected_counts')
            not_selected_counts = not_selected_df.groupby('features').size()
            not_selected_counts = not_selected_counts.to_frame(name='not_selected_counts').astype(int)

            features_summary_stats  = selected_df.groupby('features')[['feature_importances_L1_perc']].agg(['median', 'std'])
            features_summary_stats.columns = [f'{col[0]}_{col[1]}' for col in features_summary_stats.columns]

            features_summary_stats2 = united_summary_df.groupby('features')[['ci_test']].agg(['median', 'std','max'])  
            features_summary_stats2.columns = [f'{col[0]}_{col[1]}' for col in features_summary_stats2.columns]
            
            features_summary_stats3 = united_summary_df.groupby('features')[['ci_train']].agg( ['median', 'std','max'])  
            features_summary_stats3.columns = [f'{col[0]}_{col[1]}' for col in features_summary_stats3.columns]
            
            features_summary_stats4 = selected_df.groupby('features')[['auct_test']].agg(['median', 'std','max'])
            features_summary_stats4.columns = [f'{col[0]}_{col[1]}' for col in features_summary_stats4.columns]
            
            features_summary_stats5 = selected_df.groupby('features')[['auct_train']].agg(['median', 'std','max'])
            features_summary_stats5.columns = [f'{col[0]}_{col[1]}' for col in features_summary_stats5.columns]
            
            features_summary_stats6 = selected_df.groupby('features')[['#f']].agg(['median', 'std'])
            features_summary_stats6.columns = [f'{col[0]}_{col[1]}' for col in features_summary_stats6.columns]
            
            features_summary_stats = features_summary_stats.join(features_summary_stats2, lsuffix='_left', rsuffix='_right', how='outer')
            features_summary_stats = features_summary_stats.join(features_summary_stats3, lsuffix='_left', rsuffix='_right', how='outer')
            features_summary_stats = features_summary_stats.join(features_summary_stats4, lsuffix='_left', rsuffix='_right',  how='outer')
            features_summary_stats = features_summary_stats.join(features_summary_stats5, lsuffix='_left', rsuffix='_right',  how='outer')
            features_summary_stats.rename(columns={})
            
            features_summary_stats = features_summary_stats.join(features_summary_stats6, lsuffix='_left', rsuffix='_right',  how='outer')


            #adding from summary_df model run score
            stats_summary_df = features_summary_stats.join(selected_counts, lsuffix='_left', rsuffix='_right', how='outer')
            stats_summary_df = stats_summary_df.join(not_selected_counts, lsuffix='_left', rsuffix='_right', how='outer')
            stats_summary_df['not_selected_counts'] = stats_summary_df['not_selected_counts'].fillna(0).astype(int)
            stats_summary_df = stats_summary_df.join(counts_per_df, lsuffix='_left', rsuffix='_right', how='outer')
            stats_summary_df = stats_summary_df.join(agg_vectors, lsuffix='_left', rsuffix='_right', how='outer')
            stats_summary_df['selected_counts'] = stats_summary_df['selected_counts'].fillna(0)
            stats_summary_df[f'% selected'] = np.where(stats_summary_df['counts_per_feature'] != 0,
                                            (stats_summary_df['selected_counts'] / stats_summary_df['counts_per_feature']) * 100, 0).astype(int)
            # Set display options to show all columns and rows without truncation
            pd.set_option('display.max_columns', None)
            pd.set_option('display.max_rows', None)

            print(verifcation_summary_text)

            name_N_path_multiple_runs_stats = os.path.join(globals.savefig_path, multiple_runs_stats_file_name.format(token))
            stats_summary_df.to_excel(name_N_path_multiple_runs_stats, index=True, sheet_name=summary_tab_name, engine='xlsxwriter')
            print(f'writing file:{name_N_path_multiple_runs_stats}')
            workbook = load_workbook(name_N_path_multiple_runs_stats)
            errors_sheet = workbook.create_sheet(errors_tab_name)
            next_row = errors_sheet.max_row + 1
            column_number = 1  # Replace with the desired column number
            helper = f'input file to collect groups names is:{group_run_file_name}'
            errors_sheet.cell(row=next_row, column=column_number, value=helper)
            next_row = errors_sheet.max_row + 1
            errors_sheet.cell(row=next_row, column=column_number, value=verifcation_summary_text)
            next_row = errors_sheet.max_row + 2
            errors_sheet.cell(row=next_row, column=column_number, value='missing groups are:')
            for i in missing_groups:
                errors_sheet.cell(row=errors_sheet.max_row + 1, column=column_number, value=i)
                
            workbook.save(name_N_path_multiple_runs_stats)
            workbook.close()
            
    except Exception as ex:
        print(ex.__context__)
        print(''.join(traceback.format_exception(type=type(ex), value=ex, tb=ex.__traceback__)))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        current_function_name = inspect.currentframe().f_back.f_code.co_name
        print(current_function_name)
        
def features_importance_distribution(token):
    try:
        summary_file_name = os.path.join(root, folder_path, save_results_file_name.format(token))
        # Read the Excel file into a DataFrame
        summary_df_coeff = pd.read_excel(summary_file_name, sheet_name=save_results_file_name_coeff_tab_name) 
        verifcation_summary_text = 'summary file # records:{}'.format(len(summary_df_coeff))

            # Initialize lists to store feature names, coefficients, and counts
        all_feature_names = []
        all_coefficients = []

        # Iterate over rows to extract feature names and coefficients
        for i in range(0, len(summary_df_coeff), 3):
            feature_names = summary_df_coeff.iloc[i, 2:].dropna().tolist()
            try:
                coefficients = summary_df_coeff.iloc[i + 1, 2:].astype(float).dropna().tolist()
                assert len(summary_df_coeff.iloc[i]) == len(summary_df_coeff.iloc[i+1]), f"Rows {i} and {i+1} have different numbers of fields."
                # Normalize coefficients to 0-100 with absolute values
                sum = np.sum([abs(coeff) for coeff in coefficients])
                normalized_coefficients = [100*coeff/sum for coeff in coefficients]

                all_feature_names.extend(feature_names)
                all_coefficients.extend(normalized_coefficients)

            except ValueError as ve:
                print(f"Error: {ve}")
                return
                
        # Create a DataFrame with all feature names and normalized coefficients
        df_normalized = pd.DataFrame({'Feature': all_feature_names, 'Coefficient': all_coefficients})

        # Calculate the histogram bins
        bins = [-100,-50, -35, -15, -5, 5, 15, 35 ,50, 100]

        # Create a dictionary to store histograms for each name
        histograms = defaultdict(lambda: np.zeros(len(bins) - 1, dtype=int))

        # Iterate through names and numbers
        for name, number in zip(all_feature_names, all_coefficients):
            # Calculate histogram for the current number based on bins
            histogram, _ = np.histogram([number], bins=bins)
            
            # Add the histogram to the corresponding name
            histograms[name] += histogram

        # Convert the dictionary to a DataFrame
        per_f_histogram_df = pd.DataFrame(histograms)        
        
        # Create an index representing a range for each bin name
        index_range = []
        for bin_start, bin_end in zip(bins[:-1], bins[1:]):
            index_range.append(f"{bin_start}:{bin_end}")

        # Set the index to the created range
        per_f_histogram_df = per_f_histogram_df.set_index(pd.Index(index_range))
        
        
        # summary: Create a frequency table for the histogram
        histogram_table = pd.cut(df_normalized['Coefficient'], bins).value_counts().sort_index().reset_index()
        histogram_table.columns = ['Bin', 'Frequency']

        # Count how many times each feature appears
        feature_counts = df_normalized['Feature'].value_counts()

        multiple_importance_stats_file_name2 = os.path.join(globals.savefig_path, multiple_runs_stats_file_name.format(token))
        # Write results to the existing Excel file in a new tab
        with pd.ExcelWriter(multiple_importance_stats_file_name2, engine='openpyxl', mode='a') as writer:
        # Write the DataFrame to a new tab 'Histogram_Table'
            per_f_histogram_df.to_excel(writer, sheet_name=f_w_tab_name, index_label='Bins',index = True)
            histogram_table.to_excel(writer, sheet_name=f_w_tab_name+'_all', index=True)
            feature_counts.to_excel(writer, sheet_name=(f_w_tab_name+'_count'), index=True)
        #here no need for workbook.close() since the with pd.ExcelWriter do it
        print(f'writing file:{multiple_importance_stats_file_name2}')

        
    except Exception as ex:
        print(ex.__context__)
        print(''.join(traceback.format_exception(etype=type(ex), value=ex, tb=ex.__traceback__)))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        current_function_name = inspect.currentframe().f_back.f_code.co_name
        print(current_function_name)


def models_marks_distribution(token, col_name, bins):
    try:
        
        
        summary_file_name = os.path.join(root, folder_path, multiple_runs_stats_file_name.format(token))
        # Read the Excel file into a DataFrame
        multi_run_summary_df_coeff = pd.read_excel(summary_file_name, sheet_name=summary_tab_name)  
        multi_run_summary_df_coeff = multi_run_summary_df_coeff[[col_name,'features']]
        
        # Create a dictionary to store histograms for each name
        histograms = defaultdict(lambda: np.zeros(len(bins) - 1, dtype=int))

        # Iterate through names and numbers
        for f, row in multi_run_summary_df_coeff.iterrows():
            try:
                nums = row[col_name]
                numbers_string = nums.replace("nan", "float('nan')")
                numbers_list = ast.literal_eval(numbers_string)
                #print(numbers_list)
                numbers_list_float = [float(1-np.sqrt(num)) for num in numbers_list]
                f = row['features']
                # Calculate histogram for the current number based on bins
                histogram, _ = np.histogram([numbers_list_float], bins=bins)
                
                # Add the histogram to the corresponding name
                histograms[f] += histogram
            except Exception as ex:
                print(f)
                continue

        # Convert the dictionary to a DataFrame
        per_f_histogram_df = pd.DataFrame(histograms)
        
        
        # Create an index representing a range for each bin name
        index_range = []
        for bin_start, bin_end in zip(bins[:-1], bins[1:]):
            index_range.append(f"{bin_start}:{bin_end}")

        # Set the index to the created range
        per_f_histogram_df = per_f_histogram_df.set_index(pd.Index(index_range))

        sheet_name = col_name + m_w_tab_name
        multiple_importance_stats_file_name2 = os.path.join(globals.savefig_path, multiple_runs_stats_file_name.format(token))
        # Write results to the existing Excel file in a new tab
        with pd.ExcelWriter(multiple_importance_stats_file_name2, engine='openpyxl', mode='a') as writer:
        # Write the DataFrame to a new tab 'Histogram_Table'
            per_f_histogram_df.to_excel(writer, sheet_name=sheet_name, index_label='Bins',index = True)
    
        print(f'writing file:{multiple_importance_stats_file_name2}')

        
    except Exception as ex:
        print(ex.__context__)
        print(''.join(traceback.format_exception(type=type(ex), value=ex, tb=ex.__traceback__)))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        current_function_name = inspect.currentframe().f_back.f_code.co_name
        print(current_function_name)


def models_marks_distribution_of_distinct_features(token, col_name):
    try:

        summary_file_name = os.path.join(root, folder_path, save_results_file_name.format(token))
        summary_df = pd.read_excel(summary_file_name, sheet_name=summary_tab_name)

        unnamed_columns = [col for col in summary_df.columns if col.startswith('Unnamed:')]
        All_rel_f = ['features'] + unnamed_columns
        All_rel_col = ['group', col_name, 'features'] + unnamed_columns
        filtered_df = summary_df[All_rel_col]
        filtered_df = filtered_df.sort_values(col_name, ascending= False)

        distinct_f = set()
        distinct_f_counts = {}
        df_distinct_features = pd.DataFrame(columns=['group', col_name, '# distinct f', 'f list', 'added now f'])
        # Iterate through names and numbers
        for i, row in filtered_df.iterrows():
            try:
                mark = row[col_name]
                f = row[All_rel_f]
                old_distinct_f = distinct_f.copy()
                distinct_f.update(f.dropna())
                added_f = distinct_f - old_distinct_f
                long_string_f = ', '.join(distinct_f)
                long_string_f_delta = ', '.join(added_f)
                new_row = {'group': row['group'], col_name: mark, '# distinct f': len(distinct_f), 'f list': long_string_f, 'added now f': long_string_f_delta}
                df_distinct_features.loc[len(df_distinct_features)] = new_row
            except Exception as ex:
                print(f)
                print('error')
                raise

        sheet_name = col_name + "_" + m_f_inc_tab_name
        multiple_importance_stats_file_name2 = os.path.join(globals.savefig_path,
                                                            multiple_runs_stats_file_name.format(token))
        # Write results to the existing Excel file in a new tab
        with pd.ExcelWriter(multiple_importance_stats_file_name2, engine='openpyxl', mode='a') as writer:
            # Write the DataFrame to a new tab 'Histogram_Table'
            df_distinct_features.to_excel(writer, sheet_name=sheet_name, index_label='Bins', index=True)

        print(f'writing file:{multiple_importance_stats_file_name2}')


    except Exception as ex:
        print(ex.__context__)
        print(''.join(traceback.format_exception(type=type(ex), value=ex, tb=ex.__traceback__)))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        current_function_name = inspect.currentframe().f_back.f_code.co_name
        print(current_function_name)
        
def summerize_multi_runs_main(token , use_different_score_metrics = True):
        summerize(token, use_different_score_metrics)
        bins = [0.0, 0.5, 0.8, 0.83, 0.86, 0.88, 0.9, 0.92, 0.94, 0.96, 0.98, 1.0]
        print('bye bye summerize_multi_runs_main')

if __name__ == "__main__":

    token =  '2024_03_05_16_42_39'
    summerize_multi_runs_main(token)